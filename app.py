import streamlit as st
import pandas as pd
from datetime import datetime, date, timedelta
from dateutil import tz
import plotly.express as px
import plotly.graph_objects as go
import smtplib
from email.mime.text import MIMEText
import firebase_admin
from firebase_admin import credentials, firestore
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

st.set_page_config(page_title="🧊 SmartExpiry Pro", layout="wide", page_icon="🧊", initial_sidebar_state="expanded")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;900&display=swap');
* { font-family: 'Inter', -apple-system, sans-serif; }
html, body { background: #f8fafc; }
.block-container { padding: 2rem 3rem; max-width: 1920px; }
.hero-section { background: linear-gradient(135deg, #0f172a 0%, #1e3a8a 40%, #2563eb 100%); border-radius: 24px; padding: 3.5rem; color: white; margin-bottom: 3rem; box-shadow: 0 25px 60px rgba(15, 23, 42, 0.3); position: relative; overflow: hidden; }
.hero-section::before { content: ''; position: absolute; top: -50%; right: -10%; width: 600px; height: 600px; background: radial-gradient(circle, rgba(59, 130, 246, 0.1) 0%, transparent 70%); border-radius: 50%; }
.hero-content { position: relative; z-index: 1; }
.hero-content h1 { font-size: 3rem; font-weight: 900; margin: 0; letter-spacing: -1px; }
.hero-content p { font-size: 1.2rem; opacity: 0.95; margin: 1rem 0 0 0; font-weight: 500; }
.hero-stats { display: grid; grid-template-columns: repeat(4, 1fr); gap: 2rem; margin-top: 2.5rem; }
.stat-box { background: rgba(255, 255, 255, 0.1); border-radius: 12px; padding: 1.5rem; backdrop-filter: blur(10px); border: 1px solid rgba(255, 255, 255, 0.2); }
.stat-value { font-size: 2.5rem; font-weight: 900; }
.stat-label { font-size: 0.85rem; opacity: 0.85; margin-top: 0.5rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px; }
.kpi-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 1.5rem; margin-bottom: 2.5rem; }
.kpi-card { background: white; border-radius: 16px; padding: 2rem; border: 1px solid #e2e8f0; box-shadow: 0 1px 3px rgba(0, 0, 0, 0.06); transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1); position: relative; }
.kpi-card::before { content: ''; position: absolute; top: 0; left: 0; right: 0; height: 5px; border-radius: 16px 16px 0 0; background: linear-gradient(90deg, var(--kpi-color1), var(--kpi-color2)); }
.kpi-card:hover { transform: translateY(-8px) scale(1.02); box-shadow: 0 20px 40px rgba(0, 0, 0, 0.1); border-color: #cbd5e1; }
.kpi-label { font-size: 0.875rem; color: #64748b; font-weight: 700; text-transform: uppercase; letter-spacing: 0.5px; }
.kpi-value { font-size: 3.5rem; font-weight: 900; color: #0f172a; margin: 0.5rem 0; }
.kpi-delta { font-size: 0.9rem; color: #10b981; font-weight: 700; }
.kpi-subtext { font-size: 0.85rem; color: #94a3b8; margin-top: 0.5rem; }
.kpi-card:nth-child(1) { --kpi-color1: #3b82f6; --kpi-color2: #2563eb; }
.kpi-card:nth-child(2) { --kpi-color1: #dc2626; --kpi-color2: #b91c1c; }
.kpi-card:nth-child(3) { --kpi-color1: #f59e0b; --kpi-color2: #d97706; }
.task-item { background: white; border-radius: 12px; padding: 1.5rem; border-left: 5px solid #3b82f6; margin-bottom: 1rem; box-shadow: 0 1px 3px rgba(0, 0, 0, 0.05); }
.task-item:hover { box-shadow: 0 8px 16px rgba(0, 0, 0, 0.1); }
.task-item.urgent { border-left-color: #dc2626; background: #fef2f2; }
.task-item.warning { border-left-color: #f59e0b; background: #fffbeb; }
.task-product { font-size: 1.1rem; font-weight: 800; color: #0f172a; }
.task-details { font-size: 0.9rem; color: #64748b; margin-top: 0.5rem; }
.chart-container { background: white; border-radius: 16px; padding: 2rem; box-shadow: 0 1px 3px rgba(0, 0, 0, 0.06); margin-bottom: 2rem; }
.stTabs [data-baseweb="tab-list"] { border-bottom: 2px solid #e2e8f0; gap: 1rem; }
.stTabs [aria-selected="true"] { border-bottom: 3px solid #3b82f6; color: #3b82f6; }
.stButton button { border-radius: 10px; font-weight: 700; padding: 0.75rem 1.5rem; font-size: 0.95rem; letter-spacing: 0.3px; transition: all 0.2s ease; border: none; }
@media (max-width: 1024px) { .kpi-grid { grid-template-columns: 1fr; } .hero-stats { grid-template-columns: repeat(2, 1fr); } .hero-content h1 { font-size: 2rem; } }
</style>
""", unsafe_allow_html=True)

@st.cache_resource
def init_db():
    try:
        if not firebase_admin._apps:
            config = dict(st.secrets["firebase"])
            if "private_key" in config:
                pk = config["private_key"]
                if isinstance(pk, str):
                    pk = pk.replace("\\n", "\n")
                    config["private_key"] = pk
            cred = credentials.Certificate(config)
            firebase_admin.initialize_app(cred)
        return firestore.client()
    except Exception as e:
        st.error(f"❌ Erreur Firebase: {str(e)}")
        st.stop()

db = init_db()
PARIS = tz.gettz("Europe/Paris")

def days_until(exp_date) -> int:
    try:
        expiry = pd.to_datetime(exp_date).date()
        return (expiry - date.today()).days
    except:
        return 999

def stage_from_days(days: int) -> str:
    if days <= 3: return "J-3"
    elif days <= 7: return "J-7"
    elif days <= 21: return "J-21"
    else: return "OK"

def stage_config(stage: str) -> dict:
    config = {
        "J-21": {"label": "À 3 semaines", "action": "📅 Planifier une promo", "emoji": "📅", "color": "#3b82f6"},
        "J-7": {"label": "À 1 semaine", "action": "⏰ Mise en avant / -20%", "emoji": "⏰", "color": "#f59e0b"},
        "J-3": {"label": "À 3 jours", "action": "🔴 Action immédiate / -50%", "emoji": "🔴", "color": "#dc2626"},
        "OK": {"label": "> 3 semaines", "action": "✅ Stock régulier", "emoji": "✅", "color": "#10b981"}
    }
    return config.get(stage, config["OK"])

def ts_to_dt(v):
    try:
        if hasattr(v, "to_datetime"):
            return v.to_datetime()
    except:
        pass
    return v

@st.cache_data(ttl=60)
def load_lots(store_id: str) -> pd.DataFrame:
    lots = []
    try:
        for doc in db.collection("stores").document(store_id).collection("lots").stream():
            d = doc.to_dict()
            d["id"] = doc.id
            d["expiryDate"] = ts_to_dt(d.get("expiryDate"))
            lots.append(d)
    except:
        pass
    if not lots:
        return pd.DataFrame(columns=["id", "productId", "lotNumber", "quantity", "expiryDate", "location"])
    df = pd.DataFrame(lots)
    df["expiryDate"] = pd.to_datetime(df["expiryDate"])
    df["daysLeft"] = df["expiryDate"].apply(days_until)
    df["stage"] = df["daysLeft"].apply(stage_from_days)
    df = df.sort_values("expiryDate")
    return df

def get_tasks_col(store_id: str):
    return db.collection("stores").document(store_id).collection("tasks")

def ensure_tasks(store_id: str, lots_df: pd.DataFrame) -> int:
    if lots_df.empty:
        return 0
    col = get_tasks_col(store_id)
    count = 0
    for _, row in lots_df.iterrows():
        if row["stage"] == "OK":
            continue
        tid = f"TASK_{row['id']}_{row['stage']}"
        payload = {
            "id": tid, "lotId": row["id"], "productId": row.get("productId", ""), "lotNumber": row.get("lotNumber", ""),
            "stage": row["stage"], "daysLeft": int(row["daysLeft"]), "quantity": int(row.get("quantity", 0)) if pd.notna(row.get("quantity")) else 0,
            "expiryDate": row["expiryDate"].to_pydatetime(), "location": row.get("location", ""), "status": "open",
            "createdAt": firestore.SERVER_TIMESTAMP, "updatedAt": firestore.SERVER_TIMESTAMP,
        }
        col.document(tid).set(payload, merge=True)
        count += 1
    return count

@st.cache_data(ttl=60)
def load_tasks(store_id: str) -> pd.DataFrame:
    docs = list(get_tasks_col(store_id).stream())
    if not docs:
        return pd.DataFrame(columns=["id", "stage", "status"])
    records = [doc.to_dict() | {"id": doc.id} for doc in docs]
    df = pd.DataFrame(records)
    if "expiryDate" in df.columns:
        df["expiryDate"] = pd.to_datetime(df["expiryDate"].apply(lambda x: x if isinstance(x, datetime) else ts_to_dt(x)))
    return df

def update_task(store_id: str, task_id: str, status: str):
    get_tasks_col(store_id).document(task_id).set({"status": status, "updatedAt": firestore.SERVER_TIMESTAMP}, merge=True)
    st.cache_data.clear()

def send_email(subject: str, html: str) -> tuple:
    try:
        cfg = st.secrets.get("email", {})
        if not all(cfg.get(k) for k in ["host", "port", "from", "to"]):
            return False, "❌ Config email incomplète"
        msg = MIMEText(html, "html", "utf-8")
        msg["Subject"] = subject
        msg["From"] = cfg["from"]
        msg["To"] = cfg["to"]
        with smtplib.SMTP(cfg["host"], int(cfg["port"])) as s:
            if cfg.get("use_tls", True):
                s.starttls()
            if cfg.get("username") and cfg.get("password"):
                s.login(cfg["username"], cfg["password"])
            s.sendmail(msg["From"], [msg["To"]], msg.as_string())
        return True, "✅ Email envoyé avec succès"
    except Exception as e:
        return False, f"❌ Erreur email: {str(e)}"

def email_digest_html(store_id: str, tasks_df: pd.DataFrame) -> str:
    rows = []
    total_urgent = 0
    for stage in ["J-21", "J-7", "J-3"]:
        sub = tasks_df[(tasks_df["stage"] == stage) & (tasks_df["status"] == "open")]
        if sub.empty:
            continue
        if stage == "J-3":
            total_urgent = len(sub)
        cfg = stage_config(stage)
        rows.append(f"<h3 style='margin: 24px 0 12px; color: #1f2937; font-size: 18px; font-weight: 800;'>{cfg['emoji']} {cfg['label']}</h3>")
        rows.append("""<table style='width:100%;border-collapse:collapse;border:1px solid #d1d5db; margin-bottom: 20px;'><tr style='background:#f3f4f6;'><th style='padding:14px;text-align:left;border:1px solid #d1d5db;font-weight:700;'>PRODUIT</th><th style='padding:14px;text-align:left;border:1px solid #d1d5db;font-weight:700;'>LOT</th><th style='padding:14px;text-align:center;border:1px solid #d1d5db;font-weight:700;'>QTÉ</th><th style='padding:14px;text-align:center;border:1px solid #d1d5db;font-weight:700;'>DLC</th><th style='padding:14px;text-align:center;border:1px solid #d1d5db;font-weight:700;'>J</th><th style='padding:14px;text-align:left;border:1px solid #d1d5db;font-weight:700;'>RAYON</th></tr>""")
        for _, r in sub.sort_values("expiryDate").iterrows():
            exp_date = pd.to_datetime(r['expiryDate']).date().strftime('%d/%m/%Y')
            days_left = int(r.get('daysLeft',0))
            qty = int(r.get('quantity',0))
            rows.append(f"""<tr style='border-bottom:1px solid #e5e7eb;'><td style='padding:12px;border:1px solid #d1d5db;'>{r.get('productId','N/A')}</td><td style='padding:12px;border:1px solid #d1d5db;font-weight:700;color:#3b82f6;'>{r.get('lotNumber','')}</td><td style='padding:12px;border:1px solid #d1d5db;text-align:center;font-weight:700;'>{qty}</td><td style='padding:12px;border:1px solid #d1d5db;text-align:center;'>{exp_date}</td><td style='padding:12px;border:1px solid #d1d5db;text-align:center;color:{"#dc2626" if days_left <= 3 else "#f59e0b"};font-weight:900;font-size:15px;'>{days_left}j</td><td style='padding:12px;border:1px solid #d1d5db;'>{r.get('location','')}</td></tr>""")
        rows.append("</table>")
    body = "".join(rows) if rows else "<p style='text-align:center;color:#6b7280;padding:2rem;'>✅ Toutes les tâches sont à jour</p>"
    return f"""<!doctype html><html><head><meta charset="utf-8"></head><body style="font-family:'Inter', Arial, sans-serif;background:#f8fafc;margin:0;padding:20px;color:#0f172a;"><div style="max-width:850px;margin:0 auto;background:#fff;border-radius:20px;overflow:hidden;box-shadow:0 20px 60px rgba(0,0,0,0.15);"><div style="background:linear-gradient(135deg,#3b82f6 0%,#2563eb 100%);padding:40px;color:#fff;text-align:center;"><h1 style="margin:0;font-size:32px;font-weight:900;">🧊 SmartExpiry</h1><h2 style="margin:12px 0 0;font-size:20px;font-weight:700;">Digest Quotidien</h2><div style="opacity:.9;margin-top:14px;font-size:14px;">Magasin: <strong>{store_id}</strong> • {datetime.now(PARIS).strftime('%d %B %Y')}</div><div style="margin-top:20px;padding:16px;background:rgba(255,255,255,0.15);border-radius:12px;display:inline-block;"><span style="font-size:36px;font-weight:900;">{total_urgent}</span><br/><span style="font-size:13px;text-transform:uppercase;">Tâches J-3 urgentes</span></div></div><div style="padding:40px;">{body}<div style="text-align:center;margin-top:32px;"><a href="#" style="display:inline-block;background:#3b82f6;color:#fff;padding:16px 40px;border-radius:12px;text-decoration:none;font-weight:800;text-transform:uppercase;letter-spacing:0.5px;">👉 Ouvrir l'application</a></div></div><div style="background:#f1f5f9;padding:24px;text-align:center;color:#64748b;font-size:12px;border-top:1px solid #e2e8f0;"><strong>SmartExpiry Pro</strong> • Gestion FEFO<br/>Zéro perte • Marges optimisées<br/>{datetime.now(PARIS).strftime('%d/%m/%Y')}</div></div></body></html>"""

def export_to_excel(lots_df: pd.DataFrame, tasks_df: pd.DataFrame, store_id: str) -> bytes:
    output = io.BytesIO()
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Lots Urgents"
    headers = ["🏷️ PRODUIT", "📦 LOT", "📊 QTÉ", "📅 DLC", "⏱️ JOURS", "📍 RAYON", "🎯 ÉTAPE"]
    ws1.append(headers)
    header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for _, row in lots_df[lots_df["stage"] != "OK"].iterrows():
        exp_date = pd.to_datetime(row['expiryDate']).date()
        ws1.append([row.get('productId', ''), row.get('lotNumber', ''), int(row.get('quantity', 0)), exp_date, int(row['daysLeft']), row.get('location', ''), row['stage']])
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws1.column_dimensions[col].width = 16
    ws2 = wb.create_sheet("Tâches Ouvertes")
    ws2.append(headers)
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if not tasks_df.empty:
        for _, row in tasks_df[tasks_df["status"] == "open"].iterrows():
            exp_date = pd.to_datetime(row['expiryDate']).date()
            ws2.append([row.get('productId', ''), row.get('lotNumber', ''), int(row.get('quantity', 0)), exp_date, int(row.get('daysLeft', 0)), row.get('location', ''), row['stage']])
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws2.column_dimensions[col].width = 16
    wb.save(output)
    output.seek(0)
    return output.getvalue()

with st.sidebar:
    st.markdown("# 🧊 **SmartExpiry Pro**")
    st.markdown("---")
    st.markdown("### 📍 Configuration")
    store_id = st.text_input("Magasin", st.secrets.get("app", {}).get("default_store", "naturalia_nanterre"))
    st.markdown("---")
    st.markdown("### ⚡ Actions")
    if st.button("🔄 Actualiser", use_container_width=True, type="secondary"):
        st.cache_data.clear()
        st.rerun()
    st.markdown("---")
    st.markdown("""### 📖 Guide
**Workflow FEFO:**
- 📅 **J-21** : Promo
- ⏰ **J-7** : -20%
- 🔴 **J-3** : -50%""")

st.markdown(f"""
<div class="hero-section">
    <div class="hero-content">
        <h1>🧊 SmartExpiry Pro</h1>
        <p>Gestion FEFO Intelligente • Zéro Perte • Marges Optimisées</p>
        <div class="hero-stats">
            <div class="stat-box"><div class="stat-value">J-21→J-3</div><div class="stat-label">Workflow</div></div>
            <div class="stat-box"><div class="stat-value">100%</div><div class="stat-label">Auto Sync</div></div>
            <div class="stat-box"><div class="stat-value">Real-time</div><div class="stat-label">Données</div></div>
            <div class="stat-box"><div class="stat-value">{store_id}</div><div class="stat-label">Magasin</div></div>
        </div>
    </div>
</div>
""", unsafe_allow_html=True)

lots_df = load_lots(store_id)
_ = ensure_tasks(store_id, lots_df)
tasks_df = load_tasks(store_id)

open_tasks = len(tasks_df[(tasks_df["status"] == "open") & (tasks_df["stage"] != "OK")]) if not tasks_df.empty else 0
done_tasks = len(tasks_df[tasks_df["status"] == "done"]) if not tasks_df.empty else 0
urgent = len(tasks_df[(tasks_df["stage"] == "J-3") & (tasks_df["status"] == "open")]) if not tasks_df.empty else 0
pipeline = len(tasks_df[(tasks_df["stage"].isin(["J-21", "J-7"])) & (tasks_df["status"] == "open")]) if not tasks_df.empty else 0
total_qty_urgent = int(lots_df[lots_df["daysLeft"] <= 7]["quantity"].sum()) if not lots_df.empty else 0

st.markdown('<div class="kpi-grid">', unsafe_allow_html=True)
col1, col2, col3 = st.columns(3)
with col1:
    st.markdown(f'<div class="kpi-card"><div class="kpi-label">📋 Tâches Ouvertes</div><div class="kpi-value">{open_tasks}</div><div class="kpi-delta">✅ {done_tasks} terminées</div><div class="kpi-subtext">En traitement</div></div>', unsafe_allow_html=True)
with col2:
    st.markdown(f'<div class="kpi-card"><div class="kpi-label">🔴 URGENT J-3</div><div class="kpi-value">{urgent}</div><div class="kpi-delta">📦 {total_qty_urgent} unités</div><div class="kpi-subtext">Action immédiate</div></div>', unsafe_allow_html=True)
with col3:
    st.markdown(f'<div class="kpi-card"><div class="kpi-label">📊 Pipeline</div><div class="kpi-value">{pipeline}</div><div class="kpi-delta">⏰ Planifiés</div><div class="kpi-subtext">À venir</div></div>', unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)

chart_col1, chart_col2 = st.columns(2)
with chart_col1:
    st.markdown('<div class="chart-container">', unsafe_allow_html=True)
    if not lots_df.empty:
        stage_counts = lots_df[lots_df["stage"] != "OK"].groupby("stage").size().reindex(["J-3", "J-7", "J-21"], fill_value=0)
        fig = go.Figure(data=[go.Bar(x=stage_counts.index, y=stage_counts.values, marker=dict(color=['#dc2626', '#f59e0b', '#3b82f6'], line=dict(color='white', width=2)), text=stage_counts.values, textposition='outside')])
        fig.update_layout(title="📊 Distribution par Étape", height=400, showlegend=False, plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
        st.plotly_chart(fig, use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)
with chart_col2:
    st.markdown('<div class="chart-container">', unsafe_allow_html=True)
    if not lots_df.empty:
        lots_urgent = lots_df[lots_df["daysLeft"] <= 21].copy()
        fig = px.scatter(lots_urgent, x="daysLeft", y="quantity", size="quantity", color="stage", color_discrete_map={"J-3": "#dc2626", "J-7": "#f59e0b", "J-21": "#3b82f6"}, title="📈 Urgence vs Quantité")
        fig.update_layout(height=400, plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
        st.plotly_chart(fig, use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)

st.markdown("---")
tab1, tab2, tab3 = st.tabs(["📊 Dashboard Tâches", "📧 Digest Email", "📥 Rapport Excel"])

with tab1:
    st.markdown("### 📋 Gestion des Tâches par Étape")
    col1, col2, col3 = st.columns(3)
    for col_idx, stage in enumerate(["J-21", "J-7", "J-3"]):
        with [col1, col2, col3][col_idx]:
            cfg = stage_config(stage)
            st.markdown(f"#### {cfg['emoji']} {cfg['label']}")
            if tasks_df.empty:
                st.info("✅ Aucune tâche")
                continue
            sub = tasks_df[tasks_df["stage"] == stage].sort_values("expiryDate")
            if sub.empty:
                st.success("✅ Complété")
                continue
            for idx, (_, row) in enumerate(sub.iterrows()):
                if row.get("status", "open") == "done":
                    continue
                exp_date = pd.to_datetime(row['expiryDate']).date()
                days_left = int(row.get('daysLeft',0))
                qty = int(row.get('quantity',0))
                task_class = "urgent" if stage == "J-3" else "warning" if stage == "J-7" else ""
                st.markdown(f'<div class="task-item {task_class}"><div class="task-product">{row.get("productId","")}</div><div class="task-details"><strong>Lot:</strong> {row.get("lotNumber","")} • <strong>Qté:</strong> {qty} • <strong>DLC:</strong> {exp_date.strftime("%d/%m")} • <strong>Rayon:</strong> {row.get("location","")}</div></div>', unsafe_allow_html=True)
                col_a, col_b = st.columns([3, 1])
                with col_b:
                    if st.button(f"✅", key=f"done_{row['id']}_{idx}"):
                        update_task(store_id, row["id"], "done")
                        st.rerun()

with tab2:
    st.markdown("### 📧 Envoyer le Digest")
    if not lots_df.empty:
        urg = lots_df[lots_df["daysLeft"] <= 7].sort_values("expiryDate")
        if not urg.empty:
            st.info(f"📊 **{len(urg)} lots urgents** • **{int(urg['quantity'].sum())} unités**")
    if st.button("📬 Envoyer le digest", type="primary", use_container_width=True):
        with st.spinner("📧 Envoi..."):
            open_df = tasks_df[(tasks_df["status"] == "open") & (tasks_df["stage"].isin(["J-21", "J-7", "J-3"]))] if not tasks_df.empty else pd.DataFrame()
            html = email_digest_html(store_id, open_df)
            ok, msg = send_email(f"🧊 SmartExpiry Digest — {len(open_df)} tâches", html)
            if ok:
                st.success(msg)
                st.balloons()
            else:
                st.error(msg)

with tab3:
    st.markdown("### 📥 Exporter en Rapport Excel")
    if st.button("⬇️ Générer le rapport", type="primary", use_container_width=True):
        excel_data = export_to_excel(lots_df, tasks_df, store_id)
        st.download_button(label="📊 Télécharger", data=excel_data, file_name=f"smartexpiry_{store_id}_{datetime.now(PARIS).strftime('%Y%m%d_%H%M')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        st.success("✅ Rapport généré")

st.markdown("---")
st.caption(f"🧊 SmartExpiry Pro v2.0 • Dernière sync: {datetime.now(PARIS).strftime('%d/%m/%Y %H:%M:%S')}")
